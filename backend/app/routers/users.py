"""Admin user management: CRUD + Excel batch upload."""
import io

from fastapi import APIRouter, Depends, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import delete, func, select
from sqlalchemy.orm import Session

from app.database import get_db
from app.middleware.auth import invalidate_user_sessions, require_admin
from app.models import AuditLog, RequestStatus, ResourceRequest, User, UserRole
from app.schemas import BatchUploadResult, UserCreate, UserOut, UserUpdate
from app.schemas.user import (
    UserBatchDelete,
    UserBatchDeleteResult,
    UserBatchStatus,
    UserBatchStatusResult,
)
from app.services.password import hash_password

router = APIRouter(prefix="/users", tags=["users"], dependencies=[Depends(require_admin)])


def _to_out(u: User, active_resources: int = 0) -> UserOut:
    return UserOut(
        id=u.id,
        email=u.email,
        name=u.name,
        role=u.role.value,
        is_active=u.is_active,
        created_at=u.created_at,
        active_resources=active_resources,
    )


# Statuses meaning "a running instance exists on the cloud"
_RUNNING_STATUSES = [RequestStatus.approved, RequestStatus.delete_pending]


@router.get("", response_model=list[UserOut])
def list_users(db: Session = Depends(get_db)):
    users = db.scalars(select(User).order_by(User.created_at.desc())).all()
    # One grouped query: running-instance count per user
    counts = dict(
        db.execute(
            select(ResourceRequest.user_id, func.count())
            .where(ResourceRequest.status.in_(_RUNNING_STATUSES))
            .group_by(ResourceRequest.user_id)
        ).all()
    )
    return [_to_out(u, counts.get(u.id, 0)) for u in users]


@router.post("", response_model=UserOut, status_code=201)
def create_user(payload: UserCreate, db: Session = Depends(get_db)):
    if db.scalar(select(User).where(User.email == payload.email)):
        raise HTTPException(status_code=409, detail="A user with this email already exists")
    user = User(
        email=payload.email,
        name=payload.name,
        password_hash=hash_password(payload.password),
        role=UserRole.user,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return _to_out(user)


@router.put("/{user_id}", response_model=UserOut)
def update_user(user_id: int, payload: UserUpdate, db: Session = Depends(get_db)):
    user = db.get(User, user_id)
    if user is None:
        raise HTTPException(status_code=404, detail="User not found")
    if user.role == UserRole.admin and payload.is_active is False:
        raise HTTPException(status_code=400, detail="The admin account cannot be disabled")
    if payload.name is not None:
        user.name = payload.name
    if payload.password:
        user.password_hash = hash_password(payload.password)
        invalidate_user_sessions(db, user.id)
    if payload.is_active is not None:
        user.is_active = payload.is_active
        if not payload.is_active:
            invalidate_user_sessions(db, user.id)
    db.commit()
    db.refresh(user)
    return _to_out(user)


@router.post("/batch-status", response_model=UserBatchStatusResult)
def batch_set_status(payload: UserBatchStatus, db: Session = Depends(get_db)):
    """Enable or disable multiple users at once.

    Disabling is all-or-nothing: if the admin account is part of the selection,
    the whole batch is rejected.
    """
    users = db.scalars(select(User).where(User.id.in_(payload.user_ids))).all()
    if not payload.is_active and any(u.role == UserRole.admin for u in users):
        raise HTTPException(
            status_code=400,
            detail=(
                "The admin account cannot be disabled. "
                "Remove it from the selection and try again."
            ),
        )
    for user in users:
        user.is_active = payload.is_active
        if not payload.is_active:
            invalidate_user_sessions(db, user.id)
    db.commit()
    return UserBatchStatusResult(
        updated=len(users), not_found=len(payload.user_ids) - len(users)
    )


_ACTIVE_STATUSES = [RequestStatus.pending, RequestStatus.approved, RequestStatus.delete_pending]


def _deletion_blocker(db: Session, user: User) -> str | None:
    """Return a human-readable reason why this user cannot be deleted, or None."""
    if user.role == UserRole.admin:
        return "the admin account cannot be deleted"
    active = db.scalar(
        select(ResourceRequest.id)
        .where(ResourceRequest.user_id == user.id, ResourceRequest.status.in_(_ACTIVE_STATUSES))
        .limit(1)
    )
    if active is not None:
        return "has active resources or pending requests — remove/resolve them first, or disable the account"
    return None


def _purge_user(db: Session, user: User) -> None:
    """Hard-delete a user together with their finished request history.

    Removes linked audit-log entries, request rows, and sessions before the
    account itself (FK order matters).
    """
    request_ids = db.scalars(
        select(ResourceRequest.id).where(ResourceRequest.user_id == user.id)
    ).all()
    if request_ids:
        db.execute(delete(AuditLog).where(AuditLog.request_id.in_(request_ids)))
        db.execute(delete(ResourceRequest).where(ResourceRequest.id.in_(request_ids)))
    invalidate_user_sessions(db, user.id)
    db.delete(user)


@router.delete("/{user_id}", status_code=204)
def delete_user(user_id: int, db: Session = Depends(get_db)):
    user = db.get(User, user_id)
    if user is None:
        raise HTTPException(status_code=404, detail="User not found")
    blocker = _deletion_blocker(db, user)
    if blocker:
        raise HTTPException(status_code=400, detail=f"Cannot delete {user.email}: {blocker}.")
    _purge_user(db, user)
    db.commit()


@router.post("/batch-delete", response_model=UserBatchDeleteResult)
def batch_delete_users(payload: UserBatchDelete, db: Session = Depends(get_db)):
    """Delete multiple users; ineligible ones are skipped and reported."""
    deleted, skipped = 0, []
    for user_id in payload.user_ids:
        user = db.get(User, user_id)
        if user is None:
            skipped.append(f"#{user_id}: user not found")
            continue
        blocker = _deletion_blocker(db, user)
        if blocker:
            skipped.append(f"{user.email}: {blocker}")
            continue
        _purge_user(db, user)
        deleted += 1
    db.commit()
    return UserBatchDeleteResult(deleted=deleted, skipped=skipped)


@router.get("/batch-template")
def download_batch_template():
    """Provide the .xlsx template for batch user upload."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    ws.append(["Email", "Name", "Password"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="user_upload_template.xlsx"'},
    )


@router.post("/batch-upload", response_model=BatchUploadResult)
async def batch_upload(file: UploadFile, db: Session = Depends(get_db)):
    """Upsert users from an .xlsx file.

    Existing email -> overwrite name and password; new email -> append.
    """
    if not (file.filename or "").endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx file")
    try:
        wb = load_workbook(io.BytesIO(await file.read()), read_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not parse the Excel file")

    ws = wb.active
    created, updated, errors = 0, 0, []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        email, name, password = (list(row) + [None, None, None])[:3]
        if not email and not name and not password:
            continue  # skip blank rows
        if not email or not name or not password:
            errors.append(f"Row {idx}: Email, Name and Password are all required")
            continue
        email = str(email).strip().lower()
        existing = db.scalar(select(User).where(User.email == email))
        if existing:
            existing.name = str(name).strip()
            existing.password_hash = hash_password(str(password))
            invalidate_user_sessions(db, existing.id)
            updated += 1
        else:
            db.add(
                User(
                    email=email,
                    name=str(name).strip(),
                    password_hash=hash_password(str(password)),
                    role=UserRole.user,
                )
            )
            created += 1
    db.commit()
    return BatchUploadResult(created=created, updated=updated, errors=errors)
